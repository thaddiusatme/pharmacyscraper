
#!/usr/bin/env python3
"""
Pharmacy Data Processor

This script processes pharmacy JSON files from the 50-state run and generates a formatted Excel report
with independent pharmacies, including state-based color coding and summary statistics.
"""

import os
import json
import glob
import pandas as pd
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
from dataclasses import dataclass, asdict
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Define chain pharmacy identifiers
CHAIN_PHARMACIES = {
    'cvs', 'walgreens', 'rite aid', 'walmart', 'target', 'costco', 'sams club',
    'kroger', 'publix', 'safeway', 'giant', 'stop & shop', 'wegmans', 'hannaford',
    'meijer', 'hy-vee', 'albertsons', 'vons', 'pavilions', 'harris teeter',
    'food lion', 'winn-dixie', 'heb', 'ralphs', 'fry\'s', 'smith\'s', 'fred meyer',
    'qfc', 'king soopers', 'duane reade', 'riteaid', 'wal-mart', 'wal mart',
    'walmart pharmacy', 'cvs pharmacy', 'walgreens pharmacy', 'rite aid pharmacy',
    'riteaid pharmacy', 'cvs/pharmacy', 'cvs store', 'walmart neighborhood market',
    # Hospital and medical facility identifiers
    'hospital', 'medical center', 'health system', 'health center', 'clinic',
    'va pharmacy', 'veterans administration', 'veterans affairs', 'va medical',
    'military', 'army', 'navy', 'air force', 'marine', 'coast guard',
    'federal', 'government', 'state hospital', 'county hospital', 'city hospital',
    'university hospital', 'academic medical', 'teaching hospital',
    'anmc', 'anthc', 'native', 'tribal', 'indian health', 'ihs',
    'kaiser', 'permanente', 'hmo', 'health maintenance'
}

# Define styles
HEADER_FILL = PatternFill(start_color='D7E4BC', end_color='D7E4BC', fill_type='solid')
RED_FILL = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
GREEN_FILL = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
LIGHT_GREEN_FILL = PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid')
HEADER_FONT = Font(bold=True)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

@dataclass
class Pharmacy:
    """Class to hold pharmacy data with consistent field names."""
    name: str
    address: str
    city: str
    state: str
    zip_code: str
    phone: str
    website: str = ""
    confidence: float = 1.0
    business_hours: str = ""
    verified: bool = False
    is_independent: bool = True  # Assume independent by default
    
    @classmethod
    def from_dict(cls, data: Dict[str, Any]) -> Optional['Pharmacy']:
        """Create a Pharmacy instance from dictionary data."""
        try:
            # Skip chain pharmacies
            title = (data.get('title') or '').lower()
            if any(chain in title for chain in CHAIN_PHARMACIES):
                return None
                
            # Safely extract address components
            address = (data.get('address') or '').strip()
            if not address:
                return None
                
            address_parts = [p.strip() for p in address.split(',') if p.strip()]
            if len(address_parts) < 3:
                return None
                
            # First try to get state from the 'state' field if it exists
            state = (data.get('state') or '').strip()
            
            # If no state in the data, try to extract from address
            if not state and len(address_parts) >= 2:
                # Try to get state from the second-to-last part of the address
                state_part = address_parts[-2].strip()
                state_parts = state_part.split()
                
                # Check if the last part of the state part is a 2-letter state code
                if state_parts and len(state_parts[-1]) == 2 and state_parts[-1].isalpha():
                    state = state_parts[-1].upper()
                elif len(state_part) > 2:  # Might be a full state name
                    state = state_part
            
            # Get city from city field or try to extract from address
            city = (data.get('city') or '').strip()
            if not city and len(address_parts) >= 2:
                city = address_parts[-2].strip().split(',')[0].strip()
            
            # Get zip code from postalCode field or try to extract from address
            zip_code = (data.get('postalCode') or '').strip()
            if not zip_code and len(address_parts) >= 1:
                # Try to find a 5-digit number in the last part of the address
                import re
                last_part = address_parts[-1].strip()
                zip_match = re.search(r'\b(\d{5})\b', last_part)
                if zip_match:
                    zip_code = zip_match.group(1)
            
            # If we still don't have a state, try to get it from the location if available
            if not state and 'location' in data and data['location']:
                # This is a simplified approach - in a real app you'd want to use reverse geocoding
                # to get the state from the coordinates
                pass
                
            # If we still don't have a state, try to get it from the address parts
            if not state and len(address_parts) >= 2:
                # Try to find a 2-letter state code in the address
                for part in address_parts:
                    part = part.strip().upper()
                    if len(part) == 2 and part.isalpha():
                        state = part
                        break
            
            # If we still don't have a state, use the last part of the address before the zip
            if not state and len(address_parts) >= 2:
                state = address_parts[-2].strip().split()[-1].upper()
                if len(state) != 2 or not state.isalpha():
                    state = ''  # Reset if not a valid state code
            
            # Format business hours if available
            hours = []
            for day in data.get('openingHours', []):
                day_name = (day.get('day') or '')[:3]  # First 3 letters of day
                time_range = (day.get('hours') or '').replace('\u202f', ' ')  # Remove non-breaking space
                if day_name and time_range:
                    hours.append(f"{day_name} {time_range}")
            business_hours = ", ".join(hours) if hours else "Not available"
            
            # Calculate confidence score
            confidence = 0.8  # Default confidence
            if 'totalScore' in data and data['totalScore'] is not None:
                try:
                    confidence = float(data['totalScore']) / 5.0
                    confidence = max(0.0, min(1.0, confidence))  # Clamp between 0 and 1
                except (ValueError, TypeError):
                    pass
            
            return cls(
                name=(data.get('title') or '').strip(),
                address=address,
                city=city,
                state=state,
                zip_code=zip_code,
                phone=(data.get('phone') or '').strip(),
                website=(data.get('website') or '').strip(),
                confidence=confidence,
                business_hours=business_hours,
                verified=bool(data.get('verified', False))
            )
        except Exception as e:
            print(f"Error creating Pharmacy from dict: {e}")
            return None

def load_pharmacies(directory: str) -> List[Pharmacy]:
    """Load pharmacies from JSON files in the specified directory."""
    pharmacies = []
    json_files = glob.glob(os.path.join(directory, "pharmacies_*.json"))
    
    for file_path in json_files:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                for item in data:
                    if pharmacy := Pharmacy.from_dict(item):
                        pharmacies.append(pharmacy)
        except Exception as e:
            print(f"Error processing {file_path}: {e}")
    
    return pharmacies

def deduplicate_pharmacies(pharmacies: List[Pharmacy]) -> List[Pharmacy]:
    """Remove duplicate pharmacies, keeping the one with higher confidence."""
    unique_pharmacies = {}
    
    for pharm in pharmacies:
        # Create a unique key based on name and address
        key = (pharm.name.lower(), pharm.address.lower())
        
        # Keep the pharmacy with higher confidence
        if key not in unique_pharmacies or pharm.confidence > unique_pharmacies[key].confidence:
            unique_pharmacies[key] = pharm
    
    return list(unique_pharmacies.values())

def get_confidence_explanation(score: float) -> str:
    """Get a human-readable explanation of the confidence score."""
    if score >= 0.9:
        return "Very High"
    elif score >= 0.7:
        return "High"
    elif score >= 0.5:
        return "Moderate"
    else:
        return "Low"

def format_worksheet(worksheet, df: pd.DataFrame) -> None:
    """Apply formatting to a worksheet."""
    # Set column widths
    column_widths = {
        'A': 30,  # Name
        'B': 50,  # Address
        'C': 20,  # City
        'D': 15,  # State
        'E': 10,  # ZIP
        'F': 20,  # Phone
        'G': 30,  # Website
        'H': 40,  # Business Hours
        'I': 15,  # Confidence
        'J': 15,  # Confidence Level
        'K': 10,  # Verified
        'L': 15,  # Independent
    }
    
    for col_letter, width in column_widths.items():
        worksheet.column_dimensions[col_letter].width = width
    
    # Format header row
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Apply borders to all cells
    for row in worksheet.iter_rows():
        for cell in row:
            cell.border = BORDER

def format_summary_worksheet(worksheet, summary: pd.DataFrame) -> None:
    """Apply formatting to the summary worksheet."""
    # Set column widths
    worksheet.column_dimensions['A'].width = 20  # State
    worksheet.column_dimensions['B'].width = 15  # Count
    worksheet.column_dimensions['C'].width = 10  # Status
    
    # Format header row
    for cell in worksheet[1]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center')
    
    # Apply conditional formatting
    for row_idx, (_, row) in enumerate(summary.iterrows(), start=2):  # +2 for 1-based index and header
        status = worksheet[f'C{row_idx}'].value
        
        # Apply fill based on status
        if status == '<25':
            fill = RED_FILL
        elif status == '25':
            fill = LIGHT_GREEN_FILL
        else:  # >25
            fill = GREEN_FILL
        
        # Apply to all cells in the row
        for col in 'ABC':
            cell = worksheet[f'{col}{row_idx}']
            cell.fill = fill
            cell.border = BORDER

def create_excel_report(pharmacies: List[Pharmacy], output_file: str) -> None:
    """Create an Excel report with the pharmacy data."""
    if not pharmacies:
        print("No pharmacies to export.")
        return
    
    # Convert to DataFrame
    df = pd.DataFrame([asdict(pharm) for pharm in pharmacies])
    
    # Add confidence explanation
    df['confidence_level'] = df['confidence'].apply(get_confidence_explanation)
    
    # Reorder and rename columns for better readability
    column_mapping = {
        'name': 'Pharmacy Name',
        'address': 'Address',
        'city': 'City',
        'state': 'State',
        'zip_code': 'ZIP Code',
        'phone': 'Phone',
        'website': 'Website',
        'business_hours': 'Business Hours',
        'confidence': 'Confidence Score',
        'confidence_level': 'Confidence Level',
        'verified': 'Verified',
        'is_independent': 'Is Independent'
    }
    
    # Only include columns that exist in the DataFrame
    existing_columns = [col for col in column_mapping.keys() if col in df.columns]
    
    # Sort before renaming columns
    sort_columns = ['state', 'city', 'name']
    if all(col in df.columns for col in sort_columns):
        df = df.sort_values(sort_columns)
    
    # Rename columns after sorting
    df = df[existing_columns].rename(columns=column_mapping)
    
    # Create summary data - use the original column name before renaming
    if 'state' in df.columns:
        state_col = 'state'
    elif 'State' in df.columns:
        state_col = 'State'
    else:
        print("Warning: Could not find state column in DataFrame")
        return
        
    summary = df[state_col].value_counts().reset_index()
    summary.columns = ['State', 'Pharmacy Count']
    summary = summary.sort_values('State')
    
    # Add status column
    summary['Status'] = pd.cut(
        summary['Pharmacy Count'],
        bins=[-1, 24, 25, float('inf')],
        labels=['<25', '25', '>25']
    )
    
    # Create Excel writer
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write main data
        df.to_excel(writer, sheet_name='Independent Pharmacies', index=False)
        
        # Write summary data
        summary.to_excel(writer, sheet_name='State Summary', index=False)
        
        # Get workbook and worksheet objects
        workbook = writer.book
        worksheet = writer.sheets['Independent Pharmacies']
        summary_ws = writer.sheets['State Summary']
        
        # Apply formatting
        format_worksheet(worksheet, df)
        format_summary_worksheet(summary_ws, summary)
        
        # Auto-filter on the main sheet
        worksheet.auto_filter.ref = worksheet.dimensions

def main():
    import argparse
    
    parser = argparse.ArgumentParser(description='Process pharmacy data and generate Excel report')
    parser.add_argument('--input-dir', default='data/processed_50_state',
                       help='Directory containing pharmacy JSON files')
    parser.add_argument('--output-file', 
                       default=f'independent_pharmacies_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                       help='Output Excel file path')
    
    args = parser.parse_args()
    
    print(f"Loading pharmacy data from: {args.input_dir}")
    pharmacies = load_pharmacies(args.input_dir)
    print(f"Found {len(pharmacies)} pharmacies before deduplication")
    
    pharmacies = deduplicate_pharmacies(pharmacies)
    print(f"After deduplication: {len(pharmacies)} pharmacies")
    
    # Filter out pharmacies without state or with empty state
    original_count = len(pharmacies)
    
    # First, clean up state names
    for pharm in pharmacies:
        if pharm and pharm.state:
            # Clean up the state name
            state = pharm.state.strip().title()
            
            # Convert full state names to abbreviations
            state_abbreviations = {
                'alabama': 'AL', 'alaska': 'AK', 'arizona': 'AZ', 'arkansas': 'AR', 'california': 'CA',
                'colorado': 'CO', 'connecticut': 'CT', 'delaware': 'DE', 'florida': 'FL', 'georgia': 'GA',
                'hawaii': 'HI', 'idaho': 'ID', 'illinois': 'IL', 'indiana': 'IN', 'iowa': 'IA',
                'kansas': 'KS', 'kentucky': 'KY', 'louisiana': 'LA', 'maine': 'ME', 'maryland': 'MD',
                'massachusetts': 'MA', 'michigan': 'MI', 'minnesota': 'MN', 'mississippi': 'MS',
                'missouri': 'MO', 'montana': 'MT', 'nebraska': 'NE', 'nevada': 'NV', 'new hampshire': 'NH',
                'new jersey': 'NJ', 'new mexico': 'NM', 'new york': 'NY', 'north carolina': 'NC',
                'north dakota': 'ND', 'ohio': 'OH', 'oklahoma': 'OK', 'oregon': 'OR', 'pennsylvania': 'PA',
                'rhode island': 'RI', 'south carolina': 'SC', 'south dakota': 'SD', 'tennessee': 'TN',
                'texas': 'TX', 'utah': 'UT', 'vermont': 'VT', 'virginia': 'VA', 'washington': 'WA',
                'west virginia': 'WV', 'wisconsin': 'WI', 'wyoming': 'WY',
                'district of columbia': 'DC', 'washington dc': 'DC', 'puerto rico': 'PR',
                'virgin islands': 'VI', 'guam': 'GU', 'american samoa': 'AS', 'northern mariana islands': 'MP'
            }
            
            # Convert to title case for matching
            state_lower = state.lower()
            if state_lower in state_abbreviations:
                pharm.state = state_abbreviations[state_lower]
            elif len(state) == 2 and state.isalpha():
                pharm.state = state.upper()
    
    # Now filter out any remaining invalid states
    valid_states = {
        'AL', 'AK', 'AZ', 'AR', 'CA', 'CO', 'CT', 'DE', 'FL', 'GA', 'HI', 'ID', 'IL', 'IN', 'IA',
        'KS', 'KY', 'LA', 'ME', 'MD', 'MA', 'MI', 'MN', 'MS', 'MO', 'MT', 'NE', 'NV', 'NH', 'NJ',
        'NM', 'NY', 'NC', 'ND', 'OH', 'OK', 'OR', 'PA', 'RI', 'SC', 'SD', 'TN', 'TX', 'UT', 'VT',
        'VA', 'WA', 'WV', 'WI', 'WY', 'DC', 'PR', 'VI', 'GU', 'AS', 'MP'
    }
    
    # First filter by valid states
    valid_pharmacies = [p for p in pharmacies if p and p.state and p.state.upper() in valid_states]
    if removed := len(pharmacies) - len(valid_pharmacies):
        print(f"Removed {removed} pharmacies with invalid or missing state information")
    
    if not valid_pharmacies:
        print("No valid pharmacies found with state information")
        return []
        
    # Now filter out chain pharmacies
    original_count = len(valid_pharmacies)
    independent_pharmacies = []
    
    for pharm in valid_pharmacies:
        # Check if it's a chain pharmacy
        pharm_name = (pharm.name or "").lower()
        is_chain = any(chain in pharm_name for chain in CHAIN_PHARMACIES)
        pharm.is_independent = not is_chain
        
        if pharm.is_independent:
            independent_pharmacies.append(pharm)
    
    if removed := original_count - len(independent_pharmacies):
        print(f"Removed {removed} chain pharmacies")
    print(f"Found {len(independent_pharmacies)} independent pharmacies")
    
    return independent_pharmacies

def filter_chain_pharmacies(pharmacies: List[Pharmacy]) -> List[Pharmacy]:
    """Filter out chain pharmacies and return only independent ones.
    
    Args:
        pharmacies: List of Pharmacy objects to filter
        
    Returns:
        List of independent Pharmacy objects
    """
    if not pharmacies:
        return []
    
    # Define state name to abbreviation mapping
    state_name_to_abbr = {
        'alabama': 'AL', 'alaska': 'AK', 'arizona': 'AZ', 'arkansas': 'AR', 'california': 'CA',
        'colorado': 'CO', 'connecticut': 'CT', 'delaware': 'DE', 'florida': 'FL', 'georgia': 'GA',
        'hawaii': 'HI', 'idaho': 'ID', 'illinois': 'IL', 'indiana': 'IN', 'iowa': 'IA',
        'kansas': 'KS', 'kentucky': 'KY', 'louisiana': 'LA', 'maine': 'ME', 'maryland': 'MD',
        'massachusetts': 'MA', 'michigan': 'MI', 'minnesota': 'MN', 'mississippi': 'MS',
        'missouri': 'MO', 'montana': 'MT', 'nebraska': 'NE', 'nevada': 'NV', 'new hampshire': 'NH',
        'new jersey': 'NJ', 'new mexico': 'NM', 'new york': 'NY', 'north carolina': 'NC',
        'north dakota': 'ND', 'ohio': 'OH', 'oklahoma': 'OK', 'oregon': 'OR', 'pennsylvania': 'PA',
        'rhode island': 'RI', 'south carolina': 'SC', 'south dakota': 'SD', 'tennessee': 'TN',
        'texas': 'TX', 'utah': 'UT', 'vermont': 'VT', 'virginia': 'VA', 'washington': 'WA',
        'west virginia': 'WV', 'wisconsin': 'WI', 'wyoming': 'WY',
        'district of columbia': 'DC', 'washington dc': 'DC', 'puerto rico': 'PR',
        'virgin islands': 'VI', 'guam': 'GU', 'american samoa': 'AS', 'northern mariana islands': 'MP'
    }
    
    # Create a set of valid state abbreviations for quick lookup
    valid_state_abbrs = set(state_name_to_abbr.values())
    
    valid_pharmacies = []
    removed_count = 0
    
    for pharm in pharmacies:
        if not pharm or not pharm.state:
            removed_count += 1
            continue
            
        state = pharm.state.strip()
        state_upper = state.upper()
        
        # Check if it's already a valid 2-letter code
        if state_upper in valid_state_abbrs:
            pharm.state = state_upper
            valid_pharmacies.append(pharm)
            continue
            
        # Check if it's a full state name (case-insensitive)
        state_lower = state.lower()
        if state_lower in state_name_to_abbr:
            pharm.state = state_name_to_abbr[state_lower]
            valid_pharmacies.append(pharm)
            continue
            
        # If we get here, the state is not in a recognized format
        removed_count += 1
    
    if removed_count > 0:
        print(f"Removed {removed_count} pharmacies with invalid or missing state information")
    
    if not valid_pharmacies:
        print("No valid pharmacies found with state information")
        return []
        
    # Now filter out chain pharmacies
    original_count = len(valid_pharmacies)
    independent_pharmacies = []
    
    for pharm in valid_pharmacies:
        # Check if it's a chain pharmacy
        pharm_name = (pharm.name or "").lower()
        is_chain = any(chain in pharm_name for chain in CHAIN_PHARMACIES)
        pharm.is_independent = not is_chain
        
        if pharm.is_independent:
            independent_pharmacies.append(pharm)
    
    if removed := original_count - len(independent_pharmacies):
        print(f"Removed {removed} chain pharmacies")
    print(f"Found {len(independent_pharmacies)} independent pharmacies")
    
    return independent_pharmacies

def main():
    import argparse
    
    # Parse command line arguments
    parser = argparse.ArgumentParser(description='Process pharmacy data and generate Excel report')
    parser.add_argument('--input-dir', type=str, default='data/processed_50_state',
                      help='Directory containing the JSON files (default: data/processed_50_state)')
    parser.add_argument('--output-file', type=str, default=f'independent_pharmacies_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx',
                      help='Output Excel file name (default: independent_pharmacies_YYYYMMDD_HHMMSS.xlsx)')
    args = parser.parse_args()
    
    # Load and process pharmacies
    print(f"Loading pharmacy data from: {args.input_dir}")
    pharmacies = load_pharmacies(args.input_dir)
    
    if not pharmacies:
        print("No pharmacy data found. Exiting.")
        return
    
    print(f"Found {len(pharmacies)} pharmacies before filtering")
    
    # Filter out chain pharmacies
    pharmacies = filter_chain_pharmacies(pharmacies)
    
    if not pharmacies:
        print("No independent pharmacies found. Exiting.")
        return
    
    # Deduplicate pharmacies
    pharmacies = deduplicate_pharmacies(pharmacies)
    print(f"After deduplication: {len(pharmacies)} pharmacies")
    
    # Generate Excel report
    print(f"Generating Excel report: {args.output_file}")
    create_excel_report(pharmacies, args.output_file)
    
    print("\nStates with fewer than 25 pharmacies:")
    state_counts = {}
    for pharm in pharmacies:
        state_counts[pharm.state] = state_counts.get(pharm.state, 0) + 1
    
    for state, count in sorted(state_counts.items()):
        if count < 25:
            print(f"  {state}: {count} pharmacies")
    
    print(f"\nReport generated successfully: {args.output_file}")
    
    # Print states with less than 25 pharmacies
    state_counts = {}
    for pharm in pharmacies:
        if pharm and pharm.state:  # Additional safety check
            state_counts[pharm.state] = state_counts.get(pharm.state, 0) + 1
    
    low_states = {state: count for state, count in state_counts.items() if count < 25}
    if low_states:
        print("\n\033[91mStates with fewer than 25 pharmacies:")
        for state, count in sorted(low_states.items()):
            print(f"  {state}: {count} pharmacies")
        print("\033[0m")  # Reset color
    
    print(f"\nReport generated successfully: {args.output_file}")

if __name__ == "__main__":
    main()